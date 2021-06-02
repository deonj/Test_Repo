import re
import argparse
import os
from tkinter import filedialog
from tkinter import messagebox
from tkinter import *
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from datetime import date

#### added to first branch
### This would be my changes to add
### Added a new feature to first_branch

### Added a second branch
### second branch feature
### added another feature

### third branch deleted import re

### fourth branch fixed spelling and re-added import re

def get_section(lists, out_file_path, is_header, header, regex):
    
    output_list_str = []

    for item in lists:
        line_match = re.search(regex, item)
        if line_match:
            result_str = line_match.group().strip(",\"[")
            output_list_str.append(result_str)
    if len(output_list_str) != 0:  
        # write out section of script
        output = open(out_file_path, "a")
        if is_header:
            output.write("\n################ " + header + " ####################\n\n")
        for n in output_list_str:
            output.write(n + "\n")
            
def write_wb():

    wb = openpyxl.Workbook() # create a workbook
    save_path = os.getcwd() + '/output/' # path to output directory
    out_file_path = os.path.join(save_path, "output.txt") # full path to ouptut directory
    list_str = []
    row_count = 2 # keep track of last row written. start at secon row
    col_count = 1#  keep track of last column written. start at first column
    active_sheet = wb.active # select active sheet (default)
    active_sheet.title = "Main" # rename first sheet
    #write out todays date:
    today = date.today()
    # dd/mm/YY
    d1 = today.strftime("%d/%m/%Y")
    active_sheet.column_dimensions['A'].width = 15
    active_sheet.cell(column=1, row=1, value="Created on:")
    active_sheet.cell(column=1, row=2, value=d1)

    try:
        file = open(out_file_path, "r") # open the output.txt file
        list_str = file.readlines() # read file into a list
    finally:
        file.close()

    for item in list_str:
        name_match = re.search(r".+_PE_N\d", item)
        vpn_match = re.search(r"ip\svpn-instance.*", item)
        rd_match = re.search(r"route-distinguisher.*", item)
        rt_match = re.search(r"vpn-target.*", item)

        if name_match:
            sheet_name = name_match.group() # get sheet name
            wb.create_sheet(sheet_name) # create a new sheet
            active_sheet = wb[sheet_name] # change to active worksheet
            row_count = 2 # reset row counter
            col_count = 1 # reset column counter
            # Write out headers:
            active_sheet.cell(column=1, row=1, value="VPN Name")
            active_sheet.cell(column=2, row=1, value="Route Distinguisher")
            active_sheet.cell(column=3, row=1, value="Route Target")
            active_sheet.column_dimensions['A'].width = 35
            active_sheet.column_dimensions['B'].width = 20
            active_sheet.column_dimensions['C'].width = 38
           
            # change columns to autozize:
            col_style = active_sheet['A:C']# returns a tuple of tuples
            for n in col_style:
                # tuple of tuples so just need index 0
                n[0].alignment = Alignment(horizontal='center')
            # change cells in first row to bold:            
            row_style = active_sheet[1]
            for n in row_style:
                n.font = Font(bold=True)
                n.fill = PatternFill(patternType='solid', fgColor='0099CCFF')
        else:
            if vpn_match:
                item_strip = item[15:].strip() # remove the leading 'ip vpn-instance'
                col_count = 1 # reset column for every new ip vpn-instance
                active_sheet.cell(column=col_count, row=row_count, value=item_strip)
            elif rd_match:
                item_strip = item[19:].strip() # remove the leading 'route-distinguisher'
                col_count = col_count + 1
                active_sheet.cell(column=col_count, row=row_count, value=item_strip)
                col_count = col_count + 1
            elif rt_match:
                item_strip = item[10:].strip() # remove the leading 'vpn-target'
                active_sheet.cell(column=col_count, row=row_count, value=item_strip)
                row_count = row_count + 1 # increment the row only on rt
            else:
                col_count = col_count + 1
        active_sheet.delete_cols(4, 50) # remove excess columns
        
    wb.save(save_path + 'RD_RT.xlsx') # save the workbook
            
def open_file():
    
    global filename
    filename = filedialog.askopenfilename(initialdir = ".",
            title = "Select file",
            filetypes = (("text files","*.txt"),("all files","*.*")))
            
def button_click():
    try:
        get_all(filename) # parse json to output.txt
        write_wb() # output.txt to excel file
    except Exception as e:
            messagebox.showerror("Error!", e)

def about():
    messagebox.showinfo("About", 
    '''Ansible to excel parser (RDRT) v1.0
    April, 2021
    Deon Joseph''')
    
def get_all(filename):
    
    save_path = os.getcwd() + '/output/'
    out_file_path = os.path.join(save_path, "output.txt")
    rdrt_regex = r"(.+_PE_N\d|ip\svpn-instance|route-distinguisher|vpn-target).*$"
    
    try:
        file = open(filename, "r") # read the config file
        list_str = file.readlines()
        output = open(out_file_path, "w") # open an output script file
        
        #get_section(list_str, out_file_path, False, "", sysname_regex)
        get_section(list_str, out_file_path, False, "RD-RT", rdrt_regex) 
    
    finally:
        file.close()
        output.close()
        label_str.set("+++++ Finished +++++")
        messagebox.showinfo("Finished", "Check output folder for output.txt")

root = Tk()
root.title("RDRT Parser v1.0")
root.geometry("400x300")
root.resizable(0,0)
root.configure(bg="PaleGreen3")
#root.call('wm', 'iconphoto', root._w, PhotoImage(file='bin/icon.png')) # window icon

label_str = StringVar() # variable to change the label text displayed
label_str.set("Waiting for file....")
feedback_label = Label(root, textvariable=label_str, bg="PaleGreen3", fg="red")
feedback_label.pack()

frame_separator = Frame(relief=SUNKEN, borderwidth=1, bg="PaleGreen4")
frame_separator.pack(fill=BOTH, expand=True)

start_button = Button(root, bg="SeaGreen1", relief="raised", text="Start", command=button_click)
start_button.pack()

menu = Menu(root)
filemenu = Menu(menu, tearoff=0) # create a new menu
menu.add_cascade(label="File", menu=filemenu) # Menu dropdown
filemenu.add_command(label="Open...", command=open_file)# Menu item
filemenu.add_separator()
filemenu.add_command(label="Exit", command=root.quit) # Menu item

helpmenu = Menu(menu, tearoff=0) # create a new menu
menu.add_cascade(label="Help", menu=helpmenu) # Menu dropdown
helpmenu.add_command(label="About...", command=about) # Menu item
root.config(menu=menu) # display the menu

root.mainloop()

